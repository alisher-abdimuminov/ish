import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class Item:
	def __init__(
		self,
		config: dict,
		tech: str,
		data_type: str,
		band: str,
		date: str,
		time_frame: str,
	):
		self.config = config
		self.tech = tech
		self.data_type = data_type
		self.band = band
		self.date = str(date)
		self.time_frame = time_frame

		base = Path(config["REPORT_PATH"]) / "MAD" / time_frame / tech / data_type
		if tech != "LTE":
			base = base / band

		self._dashboard_dir = base / "ade_report_mail" / "collected"
		self._market_status_dir = base / "ade_report_market_status_mail" / "collected"
		self._top_offender_dir = base / "ade_report_top_offender_mail" / "collected"

	def _pick(self, d: Path):
		if not d.is_dir():
			return None
		matches = sorted(d.glob(f"*{self.date}*.csv"))
		return str(matches[-1]) if matches else None

	@property
	def dashboard(self):
		return self._pick(self._dashboard_dir)

	@property
	def market_status(self):
		return self._pick(self._market_status_dir)

	@property
	def top_offender(self):
		return self._pick(self._top_offender_dir)


class Combine:
	def __init__(self, base_path: Path, tech: str, group_name: str, items: list[Item]):
		self.tech = tech
		self.group_name = group_name
		self.items = items
		self.lte_kpi_formula = base_path / "files" / "ADE_KPI_formula_LTE.csv"
		self.nr_kpi_formula = base_path / "files" / "ADE_KPI_formula_NR.csv"
		self.df_lte = pd.read_csv(self.lte_kpi_formula)
		self.df_nr = pd.read_csv(self.nr_kpi_formula)
		self.kpi_direction_lte = dict(
			zip(self.df_lte["KPI_Name"], self.df_lte["KPI_Degradation_Direction"])
		)
		self.kpi_direction_nr = dict(
			zip(self.df_nr["KPI_Name"], self.df_nr["KPI_Degradation_Direction"])
		)

	def parse_percent_string(self, cell_value: str):
		"""Parses 'R:2:12.34%' and returns 0.1234 as float"""
		if pd.isna(cell_value) or not isinstance(cell_value, str):
			return None
		match = re.match(r"[RS]:\d+:(-?\d+\.?\d*)%", cell_value)
		if match:
			return float(match.group(1)) / 100
		return None

	def apply_conditional_formatting(self, path, sheet_name, tech):
		if tech == "LTE":
			self.kpi_direction = self.kpi_direction_lte
		elif tech == "NR":
			self.kpi_direction = self.kpi_direction_nr

		wb = load_workbook(path)
		ws = wb[sheet_name]

		red_fill = PatternFill(
			start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
		)
		green_fill = PatternFill(
			start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
		)

		header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
		kpi_col_index = header.index("kpi_name") + 1  # 1-based index
		max_row = ws.max_row
		max_col = ws.max_column

		for row in range(2, max_row + 1):
			kpi_name = ws.cell(row=row, column=kpi_col_index).value
			direction = self.kpi_direction.get(kpi_name)
			if not direction:
				continue

			for col in range(1, max_col + 1):
				cell = ws.cell(row=row, column=col)
				value = self.parse_percent_string(cell.value)
				if value is None:
					continue

				if direction == "HIGHER":
					if value > 0.1:
						cell.fill = red_fill
					elif value < -0.1:
						cell.fill = green_fill
				elif direction == "LOWER":
					if value > 0.1:
						cell.fill = green_fill
					elif value < -0.1:
						cell.fill = red_fill
		wb.save(path)

	def create_excel(self, group_name, group_info):
		output_path = f"{group_name}.xlsx"

		csvs = group_info.get("csvs", [])
		sheets = group_info.get("sheets", [])

		if not csvs or not sheets:
			print(f"[WARNING] No CSVs or sheets provided for group: {group_name}")
			return None

		written = False
		try:
			with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
				for csv, sheet in zip(csvs, sheets):
					if not csv or not sheet:
						print(
							f"[WARNING] Skipping due to missing CSV or sheet name for group: {group_name}"
						)
						continue
					try:
						df = pd.read_csv(csv)
						df.to_excel(writer, sheet_name=sheet, index=False)
						written = True
					except Exception as e:
						print(
							f"[ERROR] Failed to load/write CSV '{csv}' for sheet '{sheet}': {e}"
						)

				if not written:
					print(
						f"[ERROR] No sheets written for {group_name}. Skipping Excel file creation."
					)
					return None

			# Apply conditional formatting (outside writer block, as it may reopen the file)
			for sheet in group_info.get("format_sheets", []):
				try:
					self.apply_conditional_formatting(
						output_path, sheet, group_info.get("tech")
					)
				except Exception as e:
					print(f"[WARNING] Failed formatting sheet '{sheet}': {e}")

			print(f"[SUCCESS] Excel file created: {output_path}")
			return output_path

		except Exception as e:
			print(f"[FATAL ERROR] Failed to create Excel for {group_name}: {e}")
			return None

	def group_definitions(self):
		dashboards = [i.dashboard for i in self.items]
		top_offenders = [i.top_offender for i in self.items]
		dash_sheets = [f"MAD_Dashboard_{i.data_type}_{i.band}" for i in self.items]
		top_sheets = [f"Top_Offenders_{i.data_type}_{i.band}" for i in self.items]

		if self.tech == "LTE":
			return {
				f"MAD_1_{self.group_name}": {
					"csvs": dashboards + top_offenders,
					"sheets": dash_sheets + top_sheets,
					"format_sheets": dash_sheets,
					"tech": self.tech,
				}
			}
		else:
			statuses = [i.market_status for i in self.items]
			status_sheets = [f"Data_Status_{i.data_type}_{i.band}" for i in self.items]
			return {
				f"MAD_1_{self.group_name}": {
					"csvs": dashboards + statuses + top_offenders,
					"sheets": dash_sheets + status_sheets + top_sheets,
					"format_sheets": dash_sheets,
					"tech": self.tech,
				}
			}

	def run(self):
		for group_name, group_info in self.group_definitions().items():
			self.create_excel(group_name=group_name, group_info=group_info)
